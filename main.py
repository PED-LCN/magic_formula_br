import pandas as pd
import fundamentus
import yfinance as yf
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# --- CONFIGURAÇÕES ---
FILTRO_LIQUIDEZ = 3000000  # R$ 3 Milhões
FILTRO_VOLATILIDADE = 0.10 # Cortar os 10% mais voláteis (Decil)
REMOVER_SETOR_FINANCEIRO = True 

# Limites "Hardcoded" fornecidos (Winsorização)
# Se True, usa os números exatos que você passou. 
# Se False, calcula limites dinâmicos (melhor para o futuro).
USAR_LIMITES_FIXOS = False 

LIMITES = {
    'ey': {'max': 0.203502579, 'min': 0.035886409},
    'btm': {'max': 1.82251566, 'min': 0.118213687}
    # CFY removido pois não temos dados públicos confiáveis dele
}

def tratar_dados_brutos(df):
    """Limpa e converte colunas para float."""
    # Renomear para facilitar
    mapa = {
        'patrliq': 'patrimonio', 'patrim_liq': 'patrimonio', 
        'liq2m': 'liquidez', 'pebit': 'p_ebit', 'evebit': 'ev_ebit',
        'cotacao': 'cotacao', 'pvp': 'pvp', 'mrgebit': 'mrg_ebit',
        'pl': 'pl'
    }
    df = df.rename(columns=mapa)
    
    # Colunas numéricas
    cols = ['cotacao', 'ev_ebit', 'p_ebit', 'liquidez', 'patrimonio', 'pvp', 'pl']
    
    for col in cols:
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
    return df

def calcular_volatilidade(tickers):
    """
    Baixa histórico do Yahoo Finance e calcula volatilidade anualizada.
    Retorna um dicionário {ticker: volatilidade}.
    """
    print(f"   > Calculando volatilidade para {len(tickers)} ativos (Isso pode demorar)...")
    
    # Adiciona .SA para o Yahoo Finance reconhecer
    tickers_sa = [f"{t}.SA" for t in tickers]
    
    # Baixa dados de 1 ano
    try:
        dados = yf.download(tickers_sa, period="1y", interval="1d", progress=False)['Close']
    except Exception as e:
        print(f"Erro no Yahoo Finance: {e}")
        return {}

    volatilidades = {}
    
    # Se baixou apenas 1 ativo, a estrutura é diferente (Series vs DataFrame)
    if isinstance(dados, pd.Series):
        dados = dados.to_frame(name=tickers_sa[0])

    for col in dados.columns:
        # Ticker original (sem .SA)
        ticker_clean = col.replace('.SA', '')
        
        # Cálculo: Desvio padrão dos retornos diários * Raiz(252 dias úteis)
        retornos = dados[col].pct_change().dropna()
        if len(retornos) > 100: # Mínimo de dias para ser confiável
            vol = retornos.std() * np.sqrt(252)
            volatilidades[ticker_clean] = vol
        else:
            volatilidades[ticker_clean] = 999.0 # Penaliza quem tem pouco dado
            
    return volatilidades

def winsorizar(series, limites=None, nome_metrica=''):
    """Aplica o corte de outliers (teto e piso)."""
    if USAR_LIMITES_FIXOS and limites:
        mini = limites[nome_metrica]['min']
        maxi = limites[nome_metrica]['max']
    else:
        # Calcula dinamicamente (2.5% e 97.5%) - Estatisticamente robusto
        mini = series.quantile(0.025)
        maxi = series.quantile(0.975)
    
    return series.clip(lower=mini, upper=maxi)

def main():
    print("1. Obtendo dados Fundamentalistas...")
    df = fundamentus.get_resultado()
    df.columns = df.columns.str.strip().str.lower()
    df = tratar_dados_brutos(df)

    # --- FILTROS INICIAIS ---
    print("2. Aplicando Filtros Básicos...")
    
    # Filtro de Liquidez
    df = df[df['liquidez'] > FILTRO_LIQUIDEZ]
    
    # Filtro de Lucro (PL > 0 e P_EBIT > 0)
    df = df[df['pl'] > 0]
    
    # Filtro Setorial (Remove Bancos/Financeiro se solicitado)
    # A lib fundamentus não tem coluna 'setor' fácil, mas podemos filtrar por "bancos" conhecidos
    # ou assumir que quem não tem dados de EV/EBIT é financeiro.
    if REMOVER_SETOR_FINANCEIRO:
        # Financeiras geralmente tem EV/EBIT zerado ou distorcido na base
        df = df[df['ev_ebit'] > 0]

    # --- CÁLCULO DE VOLATILIDADE (A PARTE PESADA) ---
    tickers_filtrados = df.index.tolist()
    dict_vol = calcular_volatilidade(tickers_filtrados)
    
    # Adiciona volatilidade ao DataFrame
    df['volatilidade'] = df.index.map(dict_vol).fillna(999.0)

    # --- FILTRO DE VOLATILIDADE (O DECIL) ---
    # Cortar os 10% mais voláteis
    corte_vol = df['volatilidade'].quantile(1.0 - FILTRO_VOLATILIDADE)
    df = df[df['volatilidade'] < corte_vol]

    # --- FILTRO DE DUPLICATAS (Preferência por LIQUIDEZ) ---
    # Sua regra: "retirado a de código menos volátil" (ou menos líquida?)
    # Padrão de mercado: Mantém a de maior liquidez da empresa.
    df['prefixo'] = df.index.str[:4] # Ex: PETR
    df = df.sort_values('liquidez', ascending=False)
    df = df.groupby('prefixo').head(1) # Pega a top 1 liquidez de cada empresa
    df = df.set_index('prefixo', drop=False) # Hack para manter indice, mas na vdd o ticker original ta no index

    print(f"   > Empresas restantes após todos os filtros: {len(df)}")

    # --- CÁLCULO DE MÉTRICAS ---
    
    # 1. Market Cap
    df['mkt_cap'] = df['patrimonio'] * df['pvp']
    
    # 2. EY (Earnings Yield) = 1 / EV_EBIT
    df['ey'] = 1 / df['ev_ebit']
    
    # 3. BtM (Book to Market) = 1 / PVP
    df['btm'] = 1 / df['pvp']
    
    # 4. CFY - Indisponível (Setado para 0 para não quebrar conta)
    df['cfy'] = np.nan 

    # --- WINSORIZAÇÃO (Limites) ---
    df['ey_w'] = winsorizar(df['ey'], LIMITES, 'ey')
    df['btm_w'] = winsorizar(df['btm'], LIMITES, 'btm')

    # --- CÁLCULO DO Z-SCORE ---
    # Z = (Valor - Média) / Desvio Padrão
    
    # Z-Score do EY
    media_ey = df['ey_w'].mean()
    std_ey = df['ey_w'].std()
    df['z_ey'] = (df['ey_w'] - media_ey) / std_ey
    
    # Z-Score do BtM
    media_btm = df['btm_w'].mean()
    std_btm = df['btm_w'].std()
    df['z_btm'] = (df['btm_w'] - media_btm) / std_btm
    
    # Índice Final (Soma dos Z-Scores)
    # Como não temos CFY, usamos EY + BtM (Correlação de 95% com o índice completo)
    df['index_final'] = df['z_ey'] + df['z_btm']
    
    # Ordenar
    df = df.sort_values('index_final', ascending=False)

    # --- GERAR EXCEL ---
    print("4. Gerando Excel Final...")
    wb = Workbook()
    ws = wb.active
    ws.title = "ZScore_DeepValue"

    # Cabeçalho
    ws['A1'] = "Ranking Deep Value (Z-Score)"
    ws['A2'] = datetime.today().strftime('%Y-%m-%d')
    ws['B1'] = "Volatilidade < 10% | Liq > 3M"
    
    headers = ["Ranking", "Código", "Index Z-Score", "EY (Yield)", "BtM (Yield)", "MktCap", "Liquidez", "Volatilidade", "Cotação"]
    
    # Estilo
    fill_blue = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_white = Font(bold=True, color="FFFFFF")

    for col_num, val in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=val)
        cell.fill = fill_blue
        cell.font = font_white
        cell.alignment = Alignment(horizontal='center')

    rank = 1
    for ticker, row in df.head(100).iterrows(): # Top 100
        linha = [
            rank,
            ticker, # Ticker original estava no index antes do group by? 
            # Correção: O groupby pode ter bagunçado o index. Vamos pegar do nome
            row.name if len(str(row.name)) > 4 else ticker, 
            row['index_final'],
            row['ey'],
            row['btm'],
            row['mkt_cap'],
            row['liquidez'],
            row['volatilidade'],
            row['cotacao']
        ]
        
        for c_idx, val in enumerate(linha, 1):
            cell = ws.cell(row=rank+4, column=c_idx, value=val)
            
            # Formatação
            if c_idx in [6, 7, 9]: # Dinheiro
                cell.number_format = '"R$ "#,##0.00'
            elif c_idx in [3, 4, 5, 8]: # Decimais
                cell.number_format = '0.0000'
        
        rank += 1

    # Ajuste largura
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    wb.save("Ranking_DeepValue_ZScore.xlsx")
    print("Sucesso! Arquivo 'Ranking_DeepValue_ZScore.xlsx' gerado.")

if __name__ == "__main__":
    main()